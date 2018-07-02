#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author：John Hao
# 测试流畅度平均绘制时长的小脚本
# 功能：滑动当前页面两次，间隔1秒，收集gfxinfo并制表
# titlename：表示要测试的模块和最后文件保存的名字
# |-- 收集gfxinfo framestats详细帧数据信息
# |-- 用第三方模块openpyxl处理收集到的excel数据
# |-- 输出2秒内的120帧每一帧的绘制时间
# |-- 输出平均绘制时间

import time
import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.chart import LineChart, Reference, Series

titlelist = [
    ' UI 线程中发生的工作使其无法及时响应垂直同步信号', '应用处理输入事件所花的时间（ >2 毫秒表示处理输入事件时间长）',
    '正在运行的所有动画（ObjectAnimator、ViewPropertyAnimator 和通用转换）所需的时间',
    '完成布局和测量阶段所需的时间', '对树中的所有视图调用 View.draw() 所需的时间',
    '约 > 0.4 毫秒表示绘制了大量必须上传到 GPU 的新位图', 'GPU 工作量', '处理此帧所花的总时间', '达标线'
]
subtitlelist = [
    'IntendedVsync', 'HandleInputStart', 'AnimationStart',
    'PerformTraversalsStart', 'DrawStart', 'SyncStart',
    'IssueDrawCommandsStart', 'FrameCompleted', 'Avg'
]

# 要测试测模块名，最后文件会以该名称命名
titlename = "Feed"
print("Starting")

for j in range(1, 6):
    time.sleep(1)
    print("开始执行第" + str(j) + "遍")

    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    valueofwidth = 16
    ws.column_dimensions["A"].width = valueofwidth
    ws.column_dimensions["B"].width = valueofwidth
    ws.column_dimensions["C"].width = valueofwidth
    ws.column_dimensions["D"].width = valueofwidth
    ws.column_dimensions["E"].width = valueofwidth
    ws.column_dimensions["F"].width = valueofwidth
    ws.column_dimensions["G"].width = valueofwidth
    ws.column_dimensions["H"].width = valueofwidth
    ws.column_dimensions["I"].width = valueofwidth
    ws.column_dimensions["J"].width = valueofwidth
    ws.column_dimensions["K"].width = valueofwidth
    ws.column_dimensions["L"].width = valueofwidth
    ws.column_dimensions["M"].width = valueofwidth
    ws.column_dimensions["N"].width = valueofwidth

    # 重置所有计数器并汇总收集的统计信息
    os.popen("adb shell dumpsys gfxinfo com.google.android.tvlauncher reset")
    print("清理帧信息回到初始状态")

    # 模拟滑动页面操作
    for i in range(1, 3):
        print("执行滑动页面操作" + str(i) + "次")
        os.system("adb shell input swipe 600 600 100 100 200")
        time.sleep(1)
        os.system("adb shell input swipe 100 100 600 600 200")

    # 过滤、筛选精确的帧时间信息
    command = "adb shell dumpsys gfxinfo com.google.android.tvlauncher framestats | grep -A 120 'Flags'"
    r = os.popen(command)
    info = r.readlines()

    # 数据处理中
    print("缓存数据中......")
    for line in info:  #按行遍历
        # line = line.strip('\r\n')
        eachline = line.split(',')
        # 将行写入Excel表格
        ws.append(eachline)
        # print line

    # 新建sheet用来统计数据
    resultsheet = wb.create_sheet("result", 0)
    resultsheet.column_dimensions["A"].width = valueofwidth
    resultsheet.column_dimensions["B"].width = valueofwidth
    resultsheet.column_dimensions["C"].width = valueofwidth
    resultsheet.column_dimensions["D"].width = valueofwidth
    resultsheet.column_dimensions["E"].width = valueofwidth
    resultsheet.column_dimensions["F"].width = valueofwidth
    resultsheet.column_dimensions["G"].width = valueofwidth
    resultsheet.column_dimensions["H"].width = valueofwidth
    resultsheet.column_dimensions["I"].width = valueofwidth

    # 为结果页添加title说明
    resultsheet.append(titlelist)
    resultsheet.append(subtitlelist)
    # resultsheet.RowDimension(height = 5)

    # 填入公式，cell值由纳秒转换为毫秒
    for i in range(3, 123):
        resultsheet.cell(
            row=i,
            column=1,
            value="=data!C" + str(i - 1) + "-data!B" + str(i - 1))

    for i in range(3, 123):
        value = "=(data!G" + str(i - 1) + "-data!F" + str(i - 1)
        resultsheet.cell(row=i, column=2, value=value + ")/1000000")

    for i in range(3, 123):
        value = "=(data!H" + str(i - 1) + "-data!G" + str(i - 1)
        resultsheet.cell(row=i, column=3, value=value + ")/1000000")

    for i in range(3, 123):
        value = "=(data!I" + str(i - 1) + "-data!G" + str(i - 1)
        resultsheet.cell(row=i, column=4, value=value + ")/1000000")

    for i in range(3, 123):
        value = "=(data!K" + str(i - 1) + "-data!I" + str(i - 1)
        resultsheet.cell(row=i, column=5, value=value + ")/1000000")

    for i in range(3, 123):
        value = "=(data!L" + str(i - 1) + "-data!K" + str(i - 1)
        resultsheet.cell(row=i, column=6, value=value + ")/1000000")

    for i in range(3, 123):
        value = "=(data!L" + str(i - 1) + "-data!K" + str(i - 1)
        resultsheet.cell(row=i, column=7, value=value + ")/1000000")

    for i in range(3, 123):
        value = "=(data!N" + str(i - 1) + "-data!B" + str(i - 1)
        resultsheet.cell(row=i, column=8, value=value + ")/1000000")

    # 插入平均值16ms的列
    for i in range(3, 123):
        resultsheet.cell(row=i, column=9, value=16)

    # 插入平均Frame值
    resultsheet['J1'] = "平均值ms"
    resultsheet['J2'] = "=AVERAGEA(H3:H121)"

    # 画图准备
    chart = LineChart()
    chart.title = titlename + str(j)
    # chart.style = 5       #style都很丑，还不如默认的
    chart.y_axis.title = 'ms'
    chart.x_axis.title = 'Frame'
    chart.width = 30
    chart.height = 15

    # data选取范围
    data = Reference(resultsheet, min_col=8, min_row=2, max_col=9, max_row=122)
    chart.add_data(data, titles_from_data=True)

    # 创建图表,在B3位置插入
    resultsheet.add_chart(chart, "B3")

    #记录时间戳作为文件名
    # filename = time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time())) + ".xlsx"
    # wb.save(filename)

    #以执行名称 titlename作为文件名
    filename2 = titlename + str(j) + ".xlsx"
    wb.save(filename2)

    # 数据完毕
    print("缓存处理完毕，保存数据到本地" + str(filename2))
    time.sleep(3)
